import openpyxl
import json
import unicodedata

filepath = r'C:\Users\Trung Nguyen\Desktop\Đặt tên\Việt Danh Tính - 2A.xlsx'
output_dir = r'C:\Users\Trung Nguyen\Desktop\Antigravitiy\Orchesta assistant\viet-danh-tinh\public\data'
wb = openpyxl.load_workbook(filepath, data_only=True)

def remove_diacritics(s):
    """Remove Vietnamese diacritics using unicode decomposition"""
    # Normalize to decomposed form
    nfkd = unicodedata.normalize('NFKD', s)
    # Remove combining marks
    result = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Handle Đ/đ separately
    result = result.replace('Đ', 'D').replace('đ', 'd')
    return result

def normalize_key(s):
    return remove_diacritics(s.strip()).upper()

# === 1. Extract syllables with BOTH original and normalized keys ===
print("Extracting syllables...")
ws = wb['Tên thường Dùng']
syllables = {}
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=5):
    stt = row[0].value
    name = row[1].value
    strokes = row[2].value
    element = row[3].value
    if stt is not None and name is not None and strokes is not None:
        name_str = str(name).strip()
        entry = {
            'name': name_str,
            'strokes': int(strokes) if isinstance(strokes, (int, float)) else 0,
            'element': str(element).strip() if element else ''
        }
        upper_key = name_str.upper()
        norm_key = normalize_key(name_str)
        
        # Store with original uppercase key
        syllables[upper_key] = entry
        # Also store with normalized (no diacritics) key
        if norm_key != upper_key and norm_key not in syllables:
            syllables[norm_key] = entry

print(f"Total syllable entries: {len(syllables)}")
# Verify test case
print(f"NGUYEN lookup: {syllables.get('NGUYEN', 'NOT FOUND')}")
print(f"NGUYÊN lookup: {syllables.get('NGUYÊN', 'NOT FOUND')}")
print(f"THU lookup: {syllables.get('THU', 'NOT FOUND')}")
print(f"UYEN lookup: {syllables.get('UYEN', 'NOT FOUND')}")

with open(f'{output_dir}/syllables.json', 'w', encoding='utf-8') as f:
    json.dump(syllables, f, ensure_ascii=False)
print("Saved syllables.json")

# === 2. Fix Cục scores ===
print("\nDeriving Cục scores from luck ratings...")
cuc_m = json.load(open(f'{output_dir}/cuc_meanings.json', encoding='utf-8'))
scores_map = {}

for k, v in cuc_m.items():
    luck = v.get('luck', '').lower()
    meaning = v.get('meaning', '').lower()
    
    if luck == 'cát' or luck == 'cat':
        score = 10
    elif luck == 'hung':
        score = 2
    elif 'cát' in luck and 'hung' in luck:
        score = 5
    elif 'hung' in luck and 'cát' in luck:
        score = 5
    elif 'trong cát có hung' in luck:
        score = 4
    elif 'trong hung có cát' in luck:
        score = 6
    else:
        score = 5
    
    # Fine-tune based on meaning text
    if 'đại cát' in meaning:
        score = 10
    elif 'hung ác' in meaning or 'hung sát' in meaning:
        score = 1
    elif 'thuận lợi' in meaning or 'tốt đẹp' in meaning:
        score = 9
    
    scores_map[k] = {
        'number': int(k),
        'name': v.get('name', ''),
        'score': score
    }

print(f"Total Cục scores: {len(scores_map)}")
for i in [1, 4, 13, 22, 31, 45, 67, 81]:
    s = scores_map.get(str(i), {})
    print(f"  Cuc {i}: score={s.get('score','?')}, luck={cuc_m.get(str(i),{}).get('luck','?')}")

with open(f'{output_dir}/cuc_scores.json', 'w', encoding='utf-8') as f:
    json.dump(scores_map, f, ensure_ascii=False, indent=2)
print("Saved cuc_scores.json")

print("\n✅ Data fixed successfully!")
