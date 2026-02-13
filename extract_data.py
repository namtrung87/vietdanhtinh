import openpyxl
import json

filepath = r'C:\Users\Trung Nguyen\Desktop\Đặt tên\Việt Danh Tính - 2A.xlsx'
output_dir = r'C:\Users\Trung Nguyen\Desktop\Antigravitiy\Orchesta assistant\viet-danh-tinh\public\data'
wb = openpyxl.load_workbook(filepath, data_only=True)

# === 1. Extract syllables from 'Tên thường Dùng' ===
print("Extracting syllables...")
ws = wb['Tên thường Dùng']
syllables = {}
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=5):
    stt = row[0].value
    name = row[1].value
    strokes = row[2].value
    element = row[3].value
    if stt is not None and name is not None and strokes is not None:
        key = str(name).strip().upper()
        syllables[key] = {
            'name': str(name).strip(),
            'strokes': int(strokes) if isinstance(strokes, (int, float)) else 0,
            'element': str(element).strip() if element else ''
        }

# Also extract individual character stroke data (columns G onwards for letter breakdown)
# Characters in columns G-M for letter decomposition
print(f"Total syllables: {len(syllables)}")

with open(f'{output_dir}/syllables.json', 'w', encoding='utf-8') as f:
    json.dump(syllables, f, ensure_ascii=False, indent=2)
print(f"Saved syllables.json")

# === 2. Extract 81 Cục meanings from 'Ý Nghĩa Tứ Cục' ===
print("\nExtracting Cục meanings...")
ws4 = wb['Ý Nghĩa Tứ Cục']
cuc_meanings = {}
for row in ws4.iter_rows(min_row=7, max_row=ws4.max_row, min_col=2, max_col=7):
    num = row[0].value
    name = row[1].value
    luck = row[2].value
    alias = row[3].value
    palace = row[4].value
    meaning = row[5].value
    if num is not None and name is not None:
        num_int = int(num) if isinstance(num, (int, float)) else 0
        if num_int > 0:
            cuc_meanings[str(num_int)] = {
                'number': num_int,
                'name': str(name).strip() if name else '',
                'luck': str(luck).strip() if luck else '',
                'alias': str(alias).strip() if alias else '',
                'palace': str(palace).strip() if palace else '',
                'meaning': str(meaning).strip() if meaning else ''
            }

print(f"Total Cục meanings: {len(cuc_meanings)}")
with open(f'{output_dir}/cuc_meanings.json', 'w', encoding='utf-8') as f:
    json.dump(cuc_meanings, f, ensure_ascii=False, indent=2)
print(f"Saved cuc_meanings.json")

# === 3. Extract Cục scores from '81 Cục Việt Danh' or Sheet2 ===
print("\nExtracting Cục scores...")
ws2 = wb['Sheet2']
cuc_scores = {}
# Scores are in cols N-R, rows 20+
for row in ws2.iter_rows(min_row=20, max_row=100, min_col=14, max_col=21):
    num = row[0].value
    name_text = row[1].value
    score = row[3].value  # col Q (score)
    luck = row[5].value   # col S (luck type for Cư Cung Tốn)
    cuc_for_palace = row[6].value  # col T (cục number for that palace)
    
    if num is not None and name_text is not None:
        num_int = int(num) if isinstance(num, (int, float)) else 0
        score_val = int(score) if isinstance(score, (int, float)) else 0
        if num_int > 0:
            cuc_scores[str(num_int)] = {
                'number': num_int,
                'name': str(name_text).strip() if name_text else '',
                'score': score_val
            }

print(f"Total Cục scores: {len(cuc_scores)}")
with open(f'{output_dir}/cuc_scores.json', 'w', encoding='utf-8') as f:
    json.dump(cuc_scores, f, ensure_ascii=False, indent=2)
print(f"Saved cuc_scores.json")

# === 4. Extract Ngũ Hành element names from DT sheet ===
print("\nExtracting Ngũ Hành element groups...")
ws_dt = wb['DT']
ngu_hanh_groups = {}
# Row 6-10, col V-W contains element → name lists  
element_map = {6: 'MỘC', 7: 'HỎA', 8: 'THỔ', 9: 'KIM', 10: 'THỦY'}
for r_idx, element in element_map.items():
    for row in ws_dt.iter_rows(min_row=r_idx, max_row=r_idx, min_col=22, max_col=23):
        el = row[0].value
        names = row[1].value
        if el and names:
            ngu_hanh_groups[str(el).strip()] = str(names).strip()

print(f"Ngũ Hành groups: {len(ngu_hanh_groups)}")
with open(f'{output_dir}/ngu_hanh_groups.json', 'w', encoding='utf-8') as f:
    json.dump(ngu_hanh_groups, f, ensure_ascii=False, indent=2)
print(f"Saved ngu_hanh_groups.json")

# === 5. Extract 81 Cục detail from '81 Cục Việt Danh' ===
print("\nExtracting 81 Cục details...")
ws3 = wb['81 Cục Việt Danh']
cuc_details = {}
for i, row in enumerate(ws3.iter_rows(min_row=3, max_row=ws3.max_row, min_col=2, max_col=21), 3):
    num = row[0].value
    if num is not None and isinstance(num, (int, float)):
        num_int = int(num)
        cuc_name = str(row[1].value).strip() if row[1].value else ''
        alias = str(row[2].value).strip() if row[2].value else ''
        description = str(row[3].value).strip() if row[3].value else ''
        family = str(row[4].value).strip() if row[4].value else ''
        tinh_danh_dien = str(row[5].value).strip() if row[5].value else ''
        health = str(row[6].value).strip() if row[6].value else ''
        career = ''
        tinh_danh_phan = ''
        tinh_danh_bat = ''
        phuc_duc = ''
        # Try to get more columns if they exist
        if len(row) > 7 and row[7].value:
            career = str(row[7].value).strip()
        if len(row) > 8 and row[8].value:
            tinh_danh_phan = str(row[8].value).strip()
        if len(row) > 9 and row[9].value:
            tinh_danh_bat = str(row[9].value).strip()
        if len(row) > 10 and row[10].value:
            phuc_duc = str(row[10].value).strip()
            
        cuc_details[str(num_int)] = {
            'number': num_int,
            'cuc_name': cuc_name,
            'alias': alias,
            'description': description,
            'family': family,
            'tinh_danh_dien': tinh_danh_dien,
            'health': health,
            'career': career,
            'tinh_danh_phan': tinh_danh_phan,
            'tinh_danh_bat': tinh_danh_bat,
            'phuc_duc': phuc_duc
        }

print(f"Total 81 Cục details: {len(cuc_details)}")
with open(f'{output_dir}/cuc_details.json', 'w', encoding='utf-8') as f:
    json.dump(cuc_details, f, ensure_ascii=False, indent=2)
print(f"Saved cuc_details.json")

print("\n✅ All data extracted successfully!")
print(f"Files saved to: {output_dir}")
